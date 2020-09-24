import openpyxl
import numpy as np
wb = openpyxl.load_workbook('example.xlsx')

sheet = workbook.active
print(sheet.sheetnames)
print(sheet.title)

for rowNum in range(3, sheet.rows):
# skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    sheet.cell(row=rowNum, column=1) = (100*(produceName/177466715))

wb.save('example_copy.xlsx')

#
#    .insert_rows()
#    .delete_rows()
#    .insert_cols()
#    .delete_cols()
#
